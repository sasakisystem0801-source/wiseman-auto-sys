"""T1: CPlacementStatus / CPlacementResult dataclass 拡張テスト。
T3: resolve_xlsx (cache + scanner + 後方互換) + plan_c_placement 統合テスト。
"""

from __future__ import annotations

from pathlib import Path

from wiseman_hub.cloud.sheets import ChecklistRow
from wiseman_hub.config import ChecklistConfig, ReportStaffEntry
from wiseman_hub.pdf.checklist_c import (
    CPlacementResult,
    CPlacementStatus,
    cache_key,
    parse_multi_staff,
    plan_c_placement,
    resolve_xlsx,
    staff_choice_cache_key,
)


def _row(name: str = "テスト 太郎", staff: str = "宮下", facility: str = "事業所A") -> ChecklistRow:
    return ChecklistRow(name=name, monitoring_raw=None, staff=staff, facility=facility)


def test_status_enum_includes_needs_review() -> None:
    assert CPlacementStatus.NEEDS_REVIEW.value == "needs_review"
    # 既存ステータスは保たれている
    assert CPlacementStatus.PENDING.value == "pending"
    assert CPlacementStatus.SUCCESS.value == "success"
    assert CPlacementStatus.SKIPPED_NO_XLSX.value == "skipped_no_xlsx"


def test_result_default_fields_include_new_lists() -> None:
    """新規追加した xlsx_candidates / rejected_candidates / folder_tree がデフォルト初期化される。"""
    result = CPlacementResult(row=_row())
    assert result.xlsx_candidates == []
    assert result.rejected_candidates == {}
    assert result.folder_tree is None
    # 既存フィールドも維持
    assert result.sheet_candidates == []
    assert result.message == ""


def test_result_can_record_candidates_and_rejections() -> None:
    cand1 = Path("/x/a.xlsx")
    cand2 = Path("/x/b.xlsx")
    rejected = Path("/x/east_other.xlsx")
    result = CPlacementResult(
        row=_row(staff="木塚"),
        status=CPlacementStatus.NEEDS_REVIEW,
        xlsx_candidates=[cand1, cand2],
        rejected_candidates={rejected: "staff_token_mismatch"},
        message="複数候補",
    )
    assert result.status == CPlacementStatus.NEEDS_REVIEW
    assert result.xlsx_candidates == [cand1, cand2]
    assert result.rejected_candidates[rejected] == "staff_token_mismatch"


def test_result_can_record_folder_tree() -> None:
    """候補ゼロ時にレビュー UI へ渡すフォルダツリーが保持される。"""
    tree = {
        "name": "PT 宮下",
        "path": "\\\\Tera-station\\share\\PT 宮下",
        "is_dir": True,
        "children": [
            {
                "name": "リハ経過報告書",
                "path": "\\\\Tera-station\\share\\PT 宮下\\リハ経過報告書",
                "is_dir": True,
                "children": [],
            },
        ],
    }
    result = CPlacementResult(
        row=_row(),
        status=CPlacementStatus.NEEDS_REVIEW,
        folder_tree=tree,
        message="候補なし、フォルダから選択してください",
    )
    assert result.folder_tree is not None
    assert result.folder_tree["name"] == "PT 宮下"
    assert result.folder_tree["children"][0]["name"] == "リハ経過報告書"


# ---------- T3: resolve_xlsx ----------


def _entry_with_xlsx(tmp_path: Path) -> tuple[ReportStaffEntry, Path]:
    """テスト用 PT 宮下 fixture を作って ReportStaffEntry を返す。"""
    base = tmp_path / "PT 宮下"
    (base / "リハ経過報告書" / "令和8年").mkdir(parents=True)
    xlsx = base / "リハ経過報告書" / "令和8年" / "リハ経過報告書（宮下）3月    .xlsx"
    xlsx.write_text("")
    entry = ReportStaffEntry(
        base_dir=base,
        suggest_patterns=["リハ経過報告書/令和{era}年/リハ経過報告書*{month}月*.xlsx"],
    )
    return entry, xlsx


def test_cache_key_format() -> None:
    assert cache_key("宮下", 2026, 3) == "宮下:2026:3"


# ---------- Issue #314: parse_multi_staff / staff_choice_cache_key ----------


def test_parse_multi_staff_single() -> None:
    """単独担当者: 1 要素 list を返す (NFKC 経由でも表記は元のまま保持)。"""
    assert parse_multi_staff("小島") == ["小島"]


def test_parse_multi_staff_halfwidth_slash() -> None:
    """半角 / 区切り: 順序保持で 2 要素に分解。"""
    assert parse_multi_staff("小島/木塚") == ["小島", "木塚"]


def test_parse_multi_staff_fullwidth_slash() -> None:
    """全角 ／ 区切りも同様に分解 (実機スプレッドシート入力対応)。"""
    assert parse_multi_staff("小島／木塚") == ["小島", "木塚"]


def test_parse_multi_staff_mixed_slash() -> None:
    """半角/全角混在 (例 "小島/木塚／宮下") も問題なく 3 要素になる。"""
    assert parse_multi_staff("小島/木塚／宮下") == ["小島", "木塚", "宮下"]


def test_parse_multi_staff_empty_input() -> None:
    """空文字 / 空白のみは空 list (呼び出し側で SKIPPED_NO_STAFF 判定)。"""
    assert parse_multi_staff("") == []
    assert parse_multi_staff("   ") == []


def test_parse_multi_staff_strips_whitespace() -> None:
    """各要素の前後空白は除去 (スプレッドシート入力で頻出する " 小島 / 木塚 ")。"""
    assert parse_multi_staff(" 小島 / 木塚 ") == ["小島", "木塚"]


def test_parse_multi_staff_dedupes_normalized_duplicates() -> None:
    """normalize_lookup_key で正規化後同値の重複は 1 件に統合 (NFKC + 全角空白吸収)。

    例: "小島/小島" や "小島／ 小島" (全角空白) はどちらも 1 要素扱い。
    元表記は最初に出現したものを保持。
    """
    assert parse_multi_staff("小島/小島") == ["小島"]
    # 全角空白 + 半角 slash でも dedup
    assert parse_multi_staff("小島／　小島") == ["小島"]


def test_parse_multi_staff_keeps_original_order_with_dedupe() -> None:
    """元出現順は維持しつつ後続重複のみ除去 (UI 表示で予測可能な順序)。"""
    assert parse_multi_staff("木塚/小島/木塚") == ["木塚", "小島"]


def test_parse_multi_staff_handles_empty_parts() -> None:
    """区切り間が空 ("小島//木塚" や "/小島") でも空要素を除去して有効分のみ返す。"""
    assert parse_multi_staff("小島//木塚") == ["小島", "木塚"]
    assert parse_multi_staff("/小島") == ["小島"]
    assert parse_multi_staff("小島/") == ["小島"]


def test_staff_choice_cache_key_sorted_normalized() -> None:
    """staff_choice_cache key は normalize_lookup_key sort + | 区切り (順序非依存)。"""
    # sort は normalize 後の文字列で行う。小島 と 木塚 の Unicode 順は環境依存ではなく
    # Python の標準 codepoint 順 ("小"=0x5C0F, "木"=0x6728 → "木" > "小" だが
    # normalize_lookup_key 結果ベースで sort)
    assert staff_choice_cache_key(["小島", "木塚"], 2026, 3) == staff_choice_cache_key(
        ["木塚", "小島"], 2026, 3
    )


def test_staff_choice_cache_key_single() -> None:
    """単独担当者は ``"{staff}:{year}:{month}"`` (xlsx_path_cache と同形式)。"""
    assert staff_choice_cache_key(["小島"], 2026, 3) == "小島:2026:3"


def test_staff_choice_cache_key_empty_list_returns_empty() -> None:
    """空 list は空文字 (呼び出し側で hit 判定しない sentinel)。"""
    assert staff_choice_cache_key([], 2026, 3) == ""


def test_staff_choice_cache_key_uses_pipe_not_slash() -> None:
    """セパレータは ``|`` であって ``/`` ではない (TOML key の quote 回避)。"""
    key = staff_choice_cache_key(["小島", "木塚"], 2026, 3)
    assert "|" in key
    assert "/" not in key
    assert "／" not in key


def test_staff_choice_cache_key_includes_year_month() -> None:
    """異なる年月では別 key (xlsx_path_cache の月別キャッシュと同じ思想)。"""
    k1 = staff_choice_cache_key(["小島", "木塚"], 2026, 3)
    k2 = staff_choice_cache_key(["小島", "木塚"], 2026, 4)
    assert k1 != k2
    assert k1.endswith(":2026:3")
    assert k2.endswith(":2026:4")


def test_resolve_xlsx_cache_hit_returns_pending(tmp_path: Path) -> None:
    entry, xlsx = _entry_with_xlsx(tmp_path)
    cache = {"宮下:2026:3": str(xlsx)}
    result = resolve_xlsx("宮下", entry, 2026, 3, cache)
    assert result.status == CPlacementStatus.PENDING
    assert result.xlsx_path == xlsx


def test_resolve_xlsx_cache_stale_falls_through(tmp_path: Path) -> None:
    entry, _ = _entry_with_xlsx(tmp_path)
    cache = {"宮下:2026:3": str(tmp_path / "missing" / "stale.xlsx")}
    result = resolve_xlsx("宮下", entry, 2026, 3, cache)
    # cache stale で fall through、suggest_patterns で候補発見
    assert result.status == CPlacementStatus.NEEDS_REVIEW
    assert len(result.candidates) == 1


def test_resolve_xlsx_candidates_returns_needs_review(tmp_path: Path) -> None:
    entry, xlsx = _entry_with_xlsx(tmp_path)
    cache: dict[str, str] = {}
    result = resolve_xlsx("宮下", entry, 2026, 3, cache)
    # 候補単独でも自動確定しない（NEEDS_REVIEW）
    assert result.status == CPlacementStatus.NEEDS_REVIEW
    assert result.candidates == [xlsx]


def test_resolve_xlsx_legacy_template_fallback(tmp_path: Path) -> None:
    """suggest_patterns 空 + 旧 *_template が完全 path を生成する場合は PENDING。"""
    base = tmp_path / "PT 宮下"
    (base / "リハ経過報告書" / "令和8年").mkdir(parents=True)
    xlsx = base / "リハ経過報告書" / "令和8年" / "report.xlsx"
    xlsx.write_text("")
    entry = ReportStaffEntry(
        base_dir=base,
        suggest_patterns=[],
        year_subfolder_template="リハ経過報告書/令和{era}年",
        file_template="report.xlsx",
    )
    cache: dict[str, str] = {}
    result = resolve_xlsx("宮下", entry, 2026, 3, cache)
    assert result.status == CPlacementStatus.PENDING
    assert result.xlsx_path == xlsx


def test_resolve_xlsx_no_candidates_returns_folder_tree(tmp_path: Path) -> None:
    base = tmp_path / "PT 宮下"
    (base / "リハ経過報告書").mkdir(parents=True)
    # xlsx 不在
    entry = ReportStaffEntry(
        base_dir=base,
        suggest_patterns=["リハ経過報告書/令和{era}年/*.xlsx"],
    )
    result = resolve_xlsx("宮下", entry, 2026, 3, cache={})
    assert result.status == CPlacementStatus.NEEDS_REVIEW
    assert result.folder_tree is not None
    assert result.folder_tree["name"] == "PT 宮下"
    # 真に候補ゼロのケースは「候補なし」文言 (件数あり経路と分岐していることの担保)
    assert result.message == "候補なし、フォルダから選択してください"
    assert result.candidates == []


def test_resolve_xlsx_fallback_hits_returns_count_message(tmp_path: Path) -> None:
    """suggest_patterns 0 件 + scan_fallback hit のとき message と candidates が
    矛盾してはいけない契約を固定する (XlsxPickerDialog の候補リスト表示と
    Treeview 詳細列の整合性が崩れないため、両経路で同一文言「N 件候補あり」)。
    """
    base = tmp_path / "PT 平瀬"
    target_dir = base / "リハ経過報告書" / "令和8年"
    target_dir.mkdir(parents=True)
    # suggest_patterns には hit しないが、scan_fallback (max_depth=3 浅 walk) には hit する xlsx
    xlsx = target_dir / "新経過報告書 R8.3.xlsx"
    xlsx.write_text("")
    entry = ReportStaffEntry(
        base_dir=base,
        # 意図的に異なる pattern (例: 月名 "3月" を含むファイル名を想定) で suggest_patterns 0 件にする
        suggest_patterns=["リハ経過報告書/令和{era}年/リハ経過報告書*{month}月*.xlsx"],
    )
    result = resolve_xlsx("平瀬", entry, 2026, 3, cache={})
    assert result.status == CPlacementStatus.NEEDS_REVIEW
    assert result.candidates == [xlsx]
    assert result.message == "1 件候補あり、確認後に選択してください"


def test_resolve_xlsx_missing_base_dir_returns_skipped(tmp_path: Path) -> None:
    entry = ReportStaffEntry(
        base_dir=tmp_path / "nowhere",
        suggest_patterns=["x/y.xlsx"],
    )
    result = resolve_xlsx("宮下", entry, 2026, 3, cache={})
    assert result.status == CPlacementStatus.SKIPPED_NO_XLSX


def test_resolve_xlsx_empty_base_dir_returns_skipped() -> None:
    entry = ReportStaffEntry(base_dir=Path(""), suggest_patterns=["x.xlsx"])
    result = resolve_xlsx("宮下", entry, 2026, 3, cache={})
    assert result.status == CPlacementStatus.SKIPPED_NO_XLSX


# ---------- T3: plan_c_placement 統合 ----------


def _checklist_cfg(
    tmp_path: Path,
    *,
    cache: dict[str, str] | None = None,
    routing: dict[str, str] | None = None,
    suggest: list[str] | None = None,
) -> tuple[ChecklistConfig, Path]:
    fax_root = tmp_path / "FAX"
    fax_root.mkdir()
    base = tmp_path / "PT 宮下"
    (base / "リハ経過報告書" / "令和8年").mkdir(parents=True)
    xlsx = base / "リハ経過報告書" / "令和8年" / "リハ経過報告書（宮下）3月    .xlsx"
    xlsx.write_text("")
    entry = ReportStaffEntry(
        base_dir=base,
        suggest_patterns=suggest
        or ["リハ経過報告書/令和{era}年/リハ経過報告書*{month}月*.xlsx"],
    )
    cfg = ChecklistConfig(
        fax_root=fax_root,
        c_output_subfolder="経過報告書",
        facility_routing=routing or {"事業所A": "事業所A_FAX"},
        report_staff={"宮下": entry},
        xlsx_path_cache=cache or {},
    )
    return cfg, xlsx


def test_plan_c_placement_skipped_no_facility(tmp_path: Path) -> None:
    cfg, _ = _checklist_cfg(tmp_path, routing={})
    rows = [ChecklistRow(name="X", monitoring_raw=None, staff="宮下", facility="未登録居宅")]
    results = plan_c_placement(rows, cfg, 2026, 3)
    assert results[0].status == CPlacementStatus.SKIPPED_NO_FACILITY


def test_plan_c_placement_normalizes_facility_lookup(tmp_path: Path) -> None:
    """PR-γ v1: 全角空白入力（スプレッドシート）が半角空白登録の routing にマッチ。

    実機 Phase 3 で「介護相談支援センター　LEBEN」(全角空白) が
    「介護相談支援センター LEBEN」(半角空白) で登録された routing と
    マッチしなかった事象 (regression case) を固定する。
    """
    from wiseman_hub.utils.text_norm import normalize_lookup_key

    cfg, _ = _checklist_cfg(
        tmp_path,
        routing={
            normalize_lookup_key("介護相談支援センター LEBEN"): "LEBEN(メール)",
        },
    )
    rows = [
        ChecklistRow(
            name="X",
            monitoring_raw=None,
            staff="宮下",
            facility="介護相談支援センター　LEBEN",  # 全角空白
        )
    ]
    results = plan_c_placement(rows, cfg, 2026, 3)
    # 居宅未登録判定にならない（lookup 正規化が効いている）
    assert results[0].status != CPlacementStatus.SKIPPED_NO_FACILITY


def test_plan_c_placement_skipped_no_staff(tmp_path: Path) -> None:
    cfg, _ = _checklist_cfg(tmp_path)
    rows = [ChecklistRow(name="X", monitoring_raw=None, staff="未知担当者", facility="事業所A")]
    results = plan_c_placement(rows, cfg, 2026, 3)
    assert results[0].status == CPlacementStatus.SKIPPED_NO_STAFF


# ---------- Issue #314: plan_c_placement の複数担当者解決 (T4) ----------


def _multi_staff_cfg(tmp_path: Path) -> tuple[ChecklistConfig, Path]:
    """複数担当者テスト用 fixture。

    report_staff には "小島" / "木塚" を登録し、xlsx 解決経路は wlog "小島" のみ
    実体ファイルを持たせる ("木塚" は staff_entry はあるが xlsx 不在で skipped 経路
    に流す)。NEEDS_REVIEW_STAFF 経路の検証では xlsx 解決まで進まないので、
    fixture 単体での xlsx 実体は "小島" のみで十分。
    """
    fax_root = tmp_path / "FAX"
    fax_root.mkdir()
    base_kojima = tmp_path / "PT 小島"
    (base_kojima / "リハ経過報告書" / "令和8年").mkdir(parents=True)
    xlsx_kojima = (
        base_kojima
        / "リハ経過報告書"
        / "令和8年"
        / "リハ経過報告書（小島）3月    .xlsx"
    )
    xlsx_kojima.write_text("")
    base_kizuka = tmp_path / "PT 木塚"
    base_kizuka.mkdir()  # xlsx 実体なし
    entry_kojima = ReportStaffEntry(
        base_dir=base_kojima,
        suggest_patterns=["リハ経過報告書/令和{era}年/リハ経過報告書*{month}月*.xlsx"],
    )
    entry_kizuka = ReportStaffEntry(
        base_dir=base_kizuka,
        suggest_patterns=["リハ経過報告書/令和{era}年/*.xlsx"],
    )
    cfg = ChecklistConfig(
        fax_root=fax_root,
        c_output_subfolder="経過報告書",
        facility_routing={"事業所A": "事業所A_FAX"},
        report_staff={"小島": entry_kojima, "木塚": entry_kizuka},
    )
    return cfg, xlsx_kojima


def test_plan_c_placement_multi_staff_all_registered_returns_needs_review_staff(
    tmp_path: Path,
) -> None:
    """全員登録済の複数担当者 → NEEDS_REVIEW_STAFF、message に件数表示。

    Codex review High #4: 1 名のみでも自動確定しない。本ケースは 2 名で
    NEEDS_REVIEW_STAFF にし staff_candidates に元表記両者を入れる。
    """
    cfg, _ = _multi_staff_cfg(tmp_path)
    rows = [ChecklistRow(name="X", monitoring_raw=None, staff="小島/木塚", facility="事業所A")]
    results = plan_c_placement(rows, cfg, 2026, 3)
    assert results[0].status == CPlacementStatus.NEEDS_REVIEW_STAFF
    assert results[0].staff_candidates == ["小島", "木塚"]
    # message に件数 + 「担当者を選択」が入る
    assert "2 名" in results[0].message
    assert "担当者を選択" in results[0].message
    # 未登録なしの場合は marker が message に含まれない (xlsx 列フォーマッタ契約)
    assert "未登録あり" not in results[0].message
    # xlsx 解決前なので target_pdf / xlsx_path は None
    assert results[0].target_pdf is None
    assert results[0].xlsx_path is None


def test_plan_c_placement_multi_staff_partial_hit_shows_unregistered_names(
    tmp_path: Path,
) -> None:
    """部分 hit (一部 mapping 未登録) → NEEDS_REVIEW_STAFF、message に未登録名明示。

    Codex review High #4: 登録済 1 名のみでも自動確定しない。message に
    「未登録あり: <未登録名>」を含めて UI 表示契約 (_format_xlsx_cell が検出) を満たす。
    """
    cfg, _ = _multi_staff_cfg(tmp_path)
    # "宮下" は report_staff に存在しない (mapping 未登録)
    rows = [
        ChecklistRow(
            name="X", monitoring_raw=None, staff="小島/宮下", facility="事業所A"
        )
    ]
    results = plan_c_placement(rows, cfg, 2026, 3)
    assert results[0].status == CPlacementStatus.NEEDS_REVIEW_STAFF
    # 元表記全員 (登録済 + 未登録) が staff_candidates に入る (UI 側で disable 制御)
    assert results[0].staff_candidates == ["小島", "宮下"]
    # message に未登録名 + マーカー
    assert "未登録あり" in results[0].message
    assert "宮下" in results[0].message
    # 登録済件数 (1 名のみ登録済)
    assert "1 名のみ登録済" in results[0].message


def test_plan_c_placement_multi_staff_all_unregistered_skipped(
    tmp_path: Path,
) -> None:
    """全員 mapping 未登録 → SKIPPED_NO_STAFF (NEEDS_REVIEW_STAFF にしない)。

    既存単独経路 (SKIPPED_NO_STAFF) と整合性確保。staff_candidates は空のまま。
    """
    cfg, _ = _multi_staff_cfg(tmp_path)
    rows = [
        ChecklistRow(name="X", monitoring_raw=None, staff="未知A/未知B", facility="事業所A")
    ]
    results = plan_c_placement(rows, cfg, 2026, 3)
    assert results[0].status == CPlacementStatus.SKIPPED_NO_STAFF
    assert "全員未登録" in results[0].message
    # NEEDS_REVIEW_STAFF 用の staff_candidates は埋まらない
    assert results[0].staff_candidates == []


def test_plan_c_placement_multi_staff_cache_hit_proceeds_to_xlsx(
    tmp_path: Path,
) -> None:
    """staff_choice_cache hit → cache value (normalize_lookup_key 形式) で
    chosen staff を復元、その後通常の xlsx 解決経路に進む (Codex review High #3)。

    本テストでは xlsx 解決まで進めば候補 1 件で NEEDS_REVIEW (xlsx レビュー) に
    なる。staff レベルでは cache hit でユーザー判断 skip 済み。
    """
    cfg, xlsx_kojima = _multi_staff_cfg(tmp_path)
    # cache value は normalize_lookup_key 形式 (Codex High #1)
    from wiseman_hub.utils.text_norm import normalize_lookup_key
    cache_key_str = staff_choice_cache_key(["小島", "木塚"], 2026, 3)
    cfg_with_cache = ChecklistConfig(
        fax_root=cfg.fax_root,
        c_output_subfolder=cfg.c_output_subfolder,
        facility_routing=cfg.facility_routing,
        report_staff=cfg.report_staff,
        staff_choice_cache={cache_key_str: normalize_lookup_key("小島")},
    )
    rows = [
        ChecklistRow(name="X", monitoring_raw=None, staff="小島/木塚", facility="事業所A")
    ]
    results = plan_c_placement(rows, cfg_with_cache, 2026, 3)
    # cache hit で staff レベルの判断は skip、xlsx 解決経路に進み NEEDS_REVIEW
    # (xlsx 候補単独でも自動確定しない既存仕様、Codex review High-1 経由)
    assert results[0].status == CPlacementStatus.NEEDS_REVIEW
    assert results[0].xlsx_candidates == [xlsx_kojima]


def test_plan_c_placement_multi_staff_cache_stale_falls_through(
    tmp_path: Path,
) -> None:
    """staff_choice_cache に parsed_staffs に含まれない値が残る (stale) → NEEDS_REVIEW_STAFF。

    安全側に倒し、cache miss と同等に人間判断を求める。row.staff 自体が変わって
    過去 cache value が無効になる場面 (利用者の担当者が変更) を想定。
    """
    cfg, _ = _multi_staff_cfg(tmp_path)
    from wiseman_hub.utils.text_norm import normalize_lookup_key
    cache_key_str = staff_choice_cache_key(["小島", "木塚"], 2026, 3)
    # cache に "宮下" (parsed_staffs にも report_staff にも無い値) を残置
    cfg_with_stale = ChecklistConfig(
        fax_root=cfg.fax_root,
        c_output_subfolder=cfg.c_output_subfolder,
        facility_routing=cfg.facility_routing,
        report_staff=cfg.report_staff,
        staff_choice_cache={cache_key_str: normalize_lookup_key("宮下")},
    )
    rows = [
        ChecklistRow(name="X", monitoring_raw=None, staff="小島/木塚", facility="事業所A")
    ]
    results = plan_c_placement(rows, cfg_with_stale, 2026, 3)
    # stale cache value → 全員登録済の場合と同じく NEEDS_REVIEW_STAFF にフォールバック
    assert results[0].status == CPlacementStatus.NEEDS_REVIEW_STAFF
    assert results[0].staff_candidates == ["小島", "木塚"]


def test_plan_c_placement_single_staff_unchanged_path(tmp_path: Path) -> None:
    """単独担当者 (parse_multi_staff の len==1) は既存経路を通り regression なし。

    Issue #314 で _resolve_chosen_staff を導入したが、単独担当者の挙動は
    NEEDS_REVIEW (xlsx 候補単独でも自動確定しない既存仕様) のまま維持。
    """
    cfg, xlsx = _checklist_cfg(tmp_path)
    rows = [ChecklistRow(name="X", monitoring_raw=None, staff="宮下", facility="事業所A")]
    results = plan_c_placement(rows, cfg, 2026, 3)
    assert results[0].status == CPlacementStatus.NEEDS_REVIEW
    assert results[0].xlsx_candidates == [xlsx]
    # 単独経路では staff_candidates は埋まらない (NEEDS_REVIEW_STAFF 専用)
    assert results[0].staff_candidates == []


def test_plan_c_placement_empty_staff_skipped_no_staff(tmp_path: Path) -> None:
    """staff="" / 空白のみ → SKIPPED_NO_STAFF (既存単独経路と整合)。"""
    cfg, _ = _checklist_cfg(tmp_path)
    rows = [ChecklistRow(name="X", monitoring_raw=None, staff="", facility="事業所A")]
    results = plan_c_placement(rows, cfg, 2026, 3)
    assert results[0].status == CPlacementStatus.SKIPPED_NO_STAFF


def test_plan_c_placement_needs_review_propagates_candidates(tmp_path: Path) -> None:
    """cache miss で candidates が CPlacementResult に伝搬。"""
    cfg, xlsx = _checklist_cfg(tmp_path)
    rows = [ChecklistRow(name="X", monitoring_raw=None, staff="宮下", facility="事業所A")]
    results = plan_c_placement(rows, cfg, 2026, 3)
    assert results[0].status == CPlacementStatus.NEEDS_REVIEW
    assert results[0].xlsx_candidates == [xlsx]
    # NEEDS_REVIEW 段階では target_pdf は確定しない
    assert results[0].target_pdf is None


def test_plan_c_placement_cache_hit_sets_auto_prefix(tmp_path: Path) -> None:
    """AC-7: cache hit で確定した行は message に「自動: <basename>」prefix が入る。

    旧仕様では message が空文字で「決まっていないような表示」になっていた
    (xlsx 列なし + 詳細列空)。本 PR で xlsx 起源を Treeview から判別可能にする。
    """
    from openpyxl import Workbook

    fax_root = tmp_path / "FAX"
    fax_root.mkdir()
    base = tmp_path / "PT 宮下"
    base.mkdir()
    xlsx = base / "report.xlsx"
    wb = Workbook()
    wb.active.title = "テスト太郎"  # 利用者シートを 1 件用意
    wb.save(xlsx)
    cfg = ChecklistConfig(
        fax_root=fax_root,
        c_output_subfolder="経過報告書",
        facility_routing={"事業所A": "事業所A_FAX"},
        report_staff={
            "宮下": ReportStaffEntry(base_dir=base, suggest_patterns=["dummy"]),
        },
        xlsx_path_cache={"宮下:2026:3": str(xlsx)},  # cache hit を仕込む
    )
    rows = [
        ChecklistRow(
            name="テスト太郎", monitoring_raw=None, staff="宮下", facility="事業所A"
        )
    ]
    results = plan_c_placement(rows, cfg, 2026, 3)
    assert results[0].status == CPlacementStatus.PENDING
    assert results[0].message == "自動: report.xlsx"


def test_plan_c_placement_legacy_template_sets_auto_legacy_prefix(
    tmp_path: Path,
) -> None:
    """AC-7: legacy template 経由で確定した行は「自動: <basename> (legacy)」prefix。"""
    from openpyxl import Workbook

    fax_root = tmp_path / "FAX"
    fax_root.mkdir()
    base = tmp_path / "PT 宮下"
    (base / "リハ経過報告書" / "令和8年").mkdir(parents=True)
    xlsx = base / "リハ経過報告書" / "令和8年" / "report.xlsx"
    wb = Workbook()
    wb.active.title = "テスト太郎"
    wb.save(xlsx)
    cfg = ChecklistConfig(
        fax_root=fax_root,
        c_output_subfolder="経過報告書",
        facility_routing={"事業所A": "事業所A_FAX"},
        report_staff={
            "宮下": ReportStaffEntry(
                base_dir=base,
                suggest_patterns=[],  # 空にして legacy 経路に乗せる
                year_subfolder_template="リハ経過報告書/令和{era}年",
                file_template="report.xlsx",
            ),
        },
        xlsx_path_cache={},
    )
    rows = [
        ChecklistRow(
            name="テスト太郎", monitoring_raw=None, staff="宮下", facility="事業所A"
        )
    ]
    results = plan_c_placement(rows, cfg, 2026, 3)
    assert results[0].status == CPlacementStatus.PENDING
    assert results[0].message == "自動: report.xlsx (legacy)"


# ---------- T4: apply_xlsx_selection ----------


def _make_xlsx_with_sheet(path: Path, sheet_names: list[str]) -> Path:
    """openpyxl で実 xlsx を生成（シート名一致のテスト用）。"""
    from openpyxl import Workbook

    wb = Workbook()
    # デフォルトシートを最初の名前で置き換え
    wb.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(name)
    wb.save(path)
    return path


def test_apply_xlsx_selection_sheet_match_sets_pending(tmp_path: Path) -> None:
    from wiseman_hub.pdf.checklist_c import apply_xlsx_selection

    fax_root = tmp_path / "FAX"
    fax_root.mkdir()
    cfg = ChecklistConfig(
        fax_root=fax_root,
        c_output_subfolder="経過報告書",
        facility_routing={"事業所A": "事業所A_FAX"},
    )
    xlsx = _make_xlsx_with_sheet(tmp_path / "x.xlsx", ["テスト 太郎", "他"])
    result = CPlacementResult(
        row=_row(name="テスト 太郎", staff="宮下", facility="事業所A"),
        status=CPlacementStatus.NEEDS_REVIEW,
        xlsx_candidates=[xlsx, tmp_path / "y.xlsx"],
        folder_tree={"name": "x"},
        message="prev",
    )
    apply_xlsx_selection(result, xlsx, cfg)
    assert result.status == CPlacementStatus.PENDING
    assert result.xlsx_path == xlsx
    assert result.sheet_name == "テスト 太郎"
    assert result.target_pdf == fax_root / "事業所A_FAX" / "経過報告書" / "テスト 太郎.pdf"
    # NEEDS_REVIEW 時のフィールドはクリア
    assert result.xlsx_candidates == []
    assert result.folder_tree is None
    # PR (xlsx-visibility): 手動選択完了行は「選択: <basename>」prefix で起源を可視化。
    # (旧仕様は空文字で「決まっていないような表示」になっていた)
    assert result.message == f"選択: {xlsx.name}"


def test_apply_xlsx_selection_sheet_not_found(tmp_path: Path) -> None:
    from wiseman_hub.pdf.checklist_c import apply_xlsx_selection

    cfg = ChecklistConfig(
        fax_root=tmp_path,
        c_output_subfolder="経過報告書",
        facility_routing={"事業所A": "事業所A_FAX"},
    )
    xlsx = _make_xlsx_with_sheet(tmp_path / "x.xlsx", ["別人"])
    result = CPlacementResult(
        row=_row(name="テスト 太郎", staff="宮下", facility="事業所A"),
        status=CPlacementStatus.NEEDS_REVIEW,
    )
    apply_xlsx_selection(result, xlsx, cfg)
    assert result.status == CPlacementStatus.SKIPPED_NO_SHEET
    assert "別人" in result.sheet_candidates
    assert result.target_pdf is None


def test_apply_xlsx_selection_no_facility_routing(tmp_path: Path) -> None:
    from wiseman_hub.pdf.checklist_c import apply_xlsx_selection

    cfg = ChecklistConfig(
        fax_root=tmp_path, facility_routing={}, c_output_subfolder="経過報告書"
    )
    xlsx = _make_xlsx_with_sheet(tmp_path / "x.xlsx", ["テスト 太郎"])
    result = CPlacementResult(
        row=_row(name="テスト 太郎", staff="宮下", facility="未知居宅"),
        status=CPlacementStatus.NEEDS_REVIEW,
    )
    apply_xlsx_selection(result, xlsx, cfg)
    assert result.status == CPlacementStatus.SKIPPED_NO_FACILITY
