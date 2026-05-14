"""Google Drive 上の xlsx をダウンロードして月次シートを読み取る。

スプレッドシート連携 B/C PDF 自動配置機能（MVP）の入力層。

認証: GCP Service Account（`config.gcp.service_account_key_path` を流用）。
取得方式: Google Sheets API ではなく **Drive API v3 で xlsx をダウンロード**。
    対象ファイルは Google Sheets ネイティブではなく Excel xlsx（``application/
    vnd.openxmlformats-officedocument.spreadsheetml.sheet``）として保存されている
    ため、Sheets API は ``FAILED_PRECONDITION`` を返す。Drive API での alt=media
    ダウンロード → openpyxl での読込が唯一の経路。

xlsx の構造（チェックリスト.xlsx 仕様）:
    - シート名: ``25年3月`` ``26年4月`` 等の和暦月次タブ（最大 16 ヶ月）
    - 列: 「氏名」「モニタリング(要支援)」「担当者」「居宅」を含む
    - ヘッダー行は `\\n` を含むため正規化が必要
"""

from __future__ import annotations

import datetime as _dt
import io
import logging
import urllib.error
import urllib.request
from dataclasses import dataclass
from typing import Any

from google.auth.transport.requests import Request
from google.oauth2 import service_account
from openpyxl import load_workbook

from wiseman_hub.config import GcpConfig

logger = logging.getLogger(__name__)


_SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
]

# 列ヘッダー正規化: 改行を除去して strip + 内部空白を保持
_HEADER_ALIASES: dict[str, str] = {
    "氏名": "name",
    "ID": "id",
    "モニタリング(要支援)": "monitoring",
    "モニタリング (要支援)": "monitoring",
    "担当者": "staff",
    "居宅": "facility",
}


def _access_token(gcp: GcpConfig) -> str:
    """SA キーから Drive API 用のアクセストークンを取得する。"""
    # Issue #27 続編 G §4: service_account_key_path は Path 型、google-auth は
    # str を要求するため境界変換。
    creds = service_account.Credentials.from_service_account_file(  # type: ignore[no-untyped-call]
        str(gcp.service_account_key_path), scopes=_SCOPES
    )
    creds.refresh(Request())
    return creds.token  # type: ignore[no-any-return]


def download_xlsx(gcp: GcpConfig, file_id: str) -> bytes:
    """Drive API でファイル ID の xlsx バイト列をダウンロードする。

    HTTPError は呼び出し側で表示する想定。MVP のため retry なし。
    """
    token = _access_token(gcp)
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req) as resp:
            content: bytes = resp.read()
    except urllib.error.HTTPError as exc:
        logger.error("Drive API download failed: status=%s", exc.code)
        raise
    logger.info("Downloaded xlsx (%d bytes)", len(content))
    return content


def list_sheet_names(xlsx_bytes: bytes) -> list[str]:
    """xlsx 内の全シート名を返す。"""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", "").strip()


@dataclass(frozen=True)
class ChecklistRow:
    """スプレッドシート 1 行分の正規化データ（MVP で使う列のみ）。"""

    name: str  # 氏名（O列との結合キー）
    monitoring_raw: Any  # F列。日付 / "×" / "再開時" / 月文字列 等
    staff: str  # 担当者（小島/宮下/小林/平瀬/木塚 等）
    facility: str  # 居宅（O列）
    sheet_id: int = 0  # 後で ID 列が安定したら使う（26年3月以降）


def parse_sheet(xlsx_bytes: bytes, sheet_name: str) -> list[ChecklistRow]:
    """指定シートをパースして ChecklistRow のリストを返す。

    ヘッダー行を見つけて、必要な列だけ抽出する。空行や氏名が空の行はスキップ。
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return []

    header = [_normalize_header(c) for c in rows[0]]
    col_index: dict[str, int] = {}
    for idx, h in enumerate(header):
        key = _HEADER_ALIASES.get(h)
        if key and key not in col_index:
            col_index[key] = idx

    name_idx = col_index.get("name")
    if name_idx is None:
        raise ValueError(f"'氏名' column not found in sheet {sheet_name!r}")
    monitoring_idx = col_index.get("monitoring")
    staff_idx = col_index.get("staff")
    facility_idx = col_index.get("facility")

    def _cell(idx: int | None, row: tuple[Any, ...]) -> Any:
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def _str_cell(idx: int | None, row: tuple[Any, ...]) -> str:
        v = _cell(idx, row)
        return str(v).strip() if v is not None else ""

    result: list[ChecklistRow] = []
    for row in rows[1:]:
        if not row or len(row) <= name_idx:
            continue
        name_val = row[name_idx]
        if name_val is None:
            continue
        name = str(name_val).strip()
        if not name:
            continue
        result.append(
            ChecklistRow(
                name=name,
                monitoring_raw=_cell(monitoring_idx, row),
                staff=_str_cell(staff_idx, row),
                facility=_str_cell(facility_idx, row),
            )
        )
    return result


def is_monitoring_target(row: ChecklistRow) -> bool:
    """B 対象判定: モニタリング列が **日付（datetime or 「N月M日」形式）** の行のみ。

    除外: ``×`` / ``再開時`` / 空欄 / その他文字列。
    """
    v = row.monitoring_raw
    if v is None:
        return False
    if isinstance(v, (_dt.datetime, _dt.date)):
        return True
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return False
        # "4月27日" のような日付文字列を許容（簡易判定: "月" と "日" の両方を含む）
        if "月" in s and "日" in s:
            return True
    return False


def is_report_target(row: ChecklistRow) -> bool:
    """C 対象判定: 担当者列に値がある行（× / 空欄を除く）。"""
    s = row.staff
    if not s:
        return False
    return s not in ("×", "再開時")


def select_b_rows(rows: list[ChecklistRow]) -> list[ChecklistRow]:
    """B（モニタリング）配置対象の行を返す。"""
    return [r for r in rows if is_monitoring_target(r)]


def select_c_rows(rows: list[ChecklistRow]) -> list[ChecklistRow]:
    """C（経過報告書）配置対象の行を返す。"""
    return [r for r in rows if is_report_target(r)]
