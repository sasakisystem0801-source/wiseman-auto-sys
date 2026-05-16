"""C (経過報告書) PDF 自動配置エンジン（MVP）。

各行ごとに:
    1. 担当者から xlsx パスを解決（ReportStaffEntry の template 展開）
    2. xlsx 内の利用者シートを特定（氏名一致）
    3. Excel COM で 1 ページ目を PDF 化
    4. FAX 事業所フォルダ配下の経過報告書サブフォルダに配置

xlsx パステンプレート（{era}=令和年, {month}=月数値）:
    base_dir = ``\\\\Tera-station\\share\\PT 宮下``
    year_subfolder_template = ``リハ経過報告書\\令和{era}年``
    file_template = ``リハ経過報告書 (宮下) {month}月 .xlsx``
"""

from __future__ import annotations

import io
import logging
from collections.abc import Mapping
from dataclasses import dataclass, field
from enum import StrEnum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from wiseman_hub.cloud.sheets import ChecklistRow
from wiseman_hub.config import ChecklistConfig, ReportStaffEntry, is_path_configured
from wiseman_hub.pdf.excel_com import ExcelExporter
from wiseman_hub.pdf.staff_path_scanner import (
    build_folder_tree,
    scan_candidates,
    scan_fallback,
)
from wiseman_hub.pdf.year_folder import western_to_reiwa as _shared_western_to_reiwa
from wiseman_hub.utils.text_norm import normalize_lookup_key

logger = logging.getLogger(__name__)


class CPlacementStatus(StrEnum):
    PENDING = "pending"
    SUCCESS = "success"
    NEEDS_REVIEW = "needs_review"  # cache miss + 候補あり/なし、人間レビュー UI で選択待ち
    NEEDS_REVIEW_STAFF = "needs_review_staff"  # 担当者複数 (Issue #314): staff 選択 UI で確定待ち
    SKIPPED_NO_FACILITY = "skipped_no_facility"
    SKIPPED_NO_STAFF = "skipped_no_staff"  # 担当者マッピング未登録
    SKIPPED_NO_XLSX = "skipped_no_xlsx"
    SKIPPED_NO_SHEET = "skipped_no_sheet"  # xlsx 内に対象利用者シートなし
    SKIPPED_AMBIGUOUS_SHEET = "skipped_ambiguous_sheet"
    ERROR = "error"


@dataclass
class CPlacementResult:
    """C 配置プランの 1 行分の結果。

    NEEDS_REVIEW 時のフィールド:
        xlsx_candidates: scanner が glob 抽出した候補 xlsx の絶対パス list
        rejected_candidates: 候補から除外したパス → 除外理由（例: "staff_token_mismatch"）
        folder_tree: 候補ゼロ時に UI で表示する base_dir 配下のフォルダツリー。
            形式は ``{"name": str, "path": str, "is_dir": bool, "children": [...]}``。

    NEEDS_REVIEW_STAFF 時のフィールド (Issue #314):
        staff_candidates: parse_multi_staff で分解した担当者名 list (元表記、出現順保持)。
            StaffPickerDialog の radiobutton 選択肢として表示される。
    """

    row: ChecklistRow
    status: CPlacementStatus = CPlacementStatus.PENDING
    xlsx_path: Path | None = None
    sheet_name: str | None = None
    target_pdf: Path | None = None
    sheet_candidates: list[str] = field(default_factory=list)
    xlsx_candidates: list[Path] = field(default_factory=list)
    rejected_candidates: dict[Path, str] = field(default_factory=dict)
    folder_tree: dict[str, Any] | None = None
    staff_candidates: list[str] = field(default_factory=list)
    message: str = ""


def western_to_reiwa(year: int) -> int:
    """[deprecated] 西暦 → 令和年（2019 = R1）。

    PR-R<年>-C: ``pdf/year_folder.western_to_reiwa`` に統合 (DRY)。
    後方互換のため本関数名は維持し、内部実装を共通モジュールに委譲する。
    """
    return _shared_western_to_reiwa(year)


def cache_key(staff: str, year: int, month: int) -> str:
    """xlsx_path_cache の dict キー形式（"{staff}:{year}:{month}"）を組み立てる。"""
    return f"{staff}:{year}:{month}"


def parse_multi_staff(staff: str) -> list[str]:
    """担当列の文字列を `/` `／` 区切りで複数担当者に分解する (Issue #314)。

    スプレッドシート担当列が ``"小島/木塚"`` のような複合表記を持つ行を扱うため、
    半角 ``/`` と全角 ``／`` の両方を区切り文字として 1 度に分解する。

    順序保持 + dedupe:
        - UI 表示順序を予測可能にするため元の出現順を保持
        - 重複は normalize_lookup_key で正規化した後の値で判定 (例: "小島/小島"、
          "小島／ 小島" 等の表記揺れ込み重複も 1 件として扱う)

    Args:
        staff: 担当列の生文字列。空文字や None 相当 ("" / "  ") は ``[]`` を返す。

    Returns:
        分解後の担当者名 list (元表記、空要素除去、normalize 後重複除去済み)。
        単独担当 ("小島") は ``["小島"]`` を返す。
    """
    if not staff:
        return []
    # 全角／を半角/に正規化してから split (DRY)。NFKC では / と ／ は同一視されない
    # ため独自処理が必要。
    raw_parts = staff.replace("／", "/").split("/")
    seen_keys: set[str] = set()
    result: list[str] = []
    for part in raw_parts:
        trimmed = part.strip()
        if not trimmed:
            continue
        key = normalize_lookup_key(trimmed)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        result.append(trimmed)
    return result


def staff_choice_cache_key(staffs: list[str], year: int, month: int) -> str:
    """staff_choice_cache の dict キー形式を組み立てる (Issue #314)。

    複数担当者の組合せを順序非依存で再利用できるよう、各要素を
    normalize_lookup_key で正規化してから sort し、``|`` 区切りで連結する。
    ``|`` を採用するのは TOML key として slash ``/`` が quote 必須になり可読性を
    損なうため。

    例:
        staffs=["小島", "木塚"], 2026, 3 → "木塚|小島:2026:3" (sort 後)
        staffs=["木塚", "小島"], 2026, 3 → "木塚|小島:2026:3" (順序非依存)
        staffs=["小島"], 2026, 3 → "小島:2026:3" (単独は xlsx_path_cache と同形式)
        staffs=[] → "" (空の場合は呼び出し側で hit しない判定)

    Note:
        単独要素時の形式は ``cache_key`` (xlsx_path_cache 用) と一致するが、
        cache 自体が別 dict なのでキー衝突は発生しない。
    """
    if not staffs:
        return ""
    sorted_keys = sorted(normalize_lookup_key(s) for s in staffs)
    return f"{'|'.join(sorted_keys)}:{year}:{month}"


def resolve_xlsx_path(entry: ReportStaffEntry, year: int, month: int) -> Path:
    """[deprecated] 旧 MVP の単純 template 展開（後方互換専用）。

    新規コードからは ``resolve_xlsx`` 経由で cache + scanner を使うこと。
    suggest_patterns 空 + 旧 *_template フィールドが両方埋まっているときの
    フォールバック用に残置している。

    template が空のときは ``base_dir`` を返すだけになるが、呼び出し側で
    ``exists()`` 失敗扱いになるので機能はしない。
    """
    era = western_to_reiwa(year)
    # Issue #27 続編 G Phase 3b: entry.base_dir は Path 型に移行済 (重複ラップ除去)。
    base = entry.base_dir
    year_sub = entry.year_subfolder_template.format(era=era, month=month)
    fname = entry.file_template.format(era=era, month=month)
    return base / year_sub / fname


@dataclass
class ResolveResult:
    """resolve_xlsx の戻り値。

    呼び出し側 ``plan_c_placement`` が CPlacementResult に統合する。
    status は CPlacementStatus の以下のサブセットのみを返す:
        - PENDING: cache hit でパス確定（後段でシート検査）
        - NEEDS_REVIEW: 候補抽出に成功（複数 or 単独）、人間レビュー UI で選択待ち
        - SKIPPED_NO_XLSX: 候補ゼロ + folder_tree も組めない（base_dir 不在等）
    """

    status: CPlacementStatus
    xlsx_path: Path | None = None
    candidates: list[Path] = field(default_factory=list)
    folder_tree: dict[str, Any] | None = None
    message: str = ""


def resolve_xlsx(
    staff: str,
    entry: ReportStaffEntry,
    year: int,
    month: int,
    cache: Mapping[str, str],
) -> ResolveResult:
    """担当者ごとの xlsx パス解決。cache hit → PENDING / miss → NEEDS_REVIEW。

    解決順序:
        1. cache hit でファイル存在 → PENDING（自動確定）
        2. cache stale（path 不在） → fall through、再 scan
        3. suggest_patterns で候補抽出 → 単独/複数とも NEEDS_REVIEW（自動確定しない）
        4. 候補ゼロ + 旧 *_template があれば legacy fallback で path 試行
        5. 候補ゼロ → フォルダツリー + scan_fallback で NEEDS_REVIEW
        6. base_dir 不在 → SKIPPED_NO_XLSX

    自動確定するのは「cache hit」のみ。これは過去に人間が UI で選択 +
    「記憶する」を確定した path であり、deterministic な根拠を持つ。
    """
    key = cache_key(staff, year, month)
    cached_str = cache.get(key)
    if cached_str:
        cached_path = Path(cached_str)
        if cached_path.exists():
            return ResolveResult(
                status=CPlacementStatus.PENDING,
                xlsx_path=cached_path,
            )
        # cache stale
        logger.info("cache stale for %s, re-scanning", key)

    # suggest_patterns で候補絞り込み
    candidates = scan_candidates(entry, year, month)
    if candidates:
        return ResolveResult(
            status=CPlacementStatus.NEEDS_REVIEW,
            candidates=candidates,
            message=f"{len(candidates)} 件候補あり、確認後に選択してください",
        )

    # 後方互換: suggest_patterns 空かつ legacy template が埋まっている場合
    if not entry.suggest_patterns and entry.year_subfolder_template and entry.file_template:
        legacy = resolve_xlsx_path(entry, year, month)
        if legacy.exists():
            return ResolveResult(
                status=CPlacementStatus.PENDING,
                xlsx_path=legacy,
                # PR (xlsx-visibility): 「自動: <basename> (legacy)」で legacy 経路を明示。
                message=f"自動: {legacy.name} (legacy)",
            )

    # フォールバック: base_dir 配下を浅く scan + folder_tree 提示
    # Issue #27 続編 G Phase 3b: entry.base_dir は Path 型、is_path_configured で sentinel 判定。
    if not is_path_configured(entry.base_dir) or not entry.base_dir.exists():
        return ResolveResult(
            status=CPlacementStatus.SKIPPED_NO_XLSX,
            message=f"base_dir 不在または未設定: {entry.base_dir if is_path_configured(entry.base_dir) else '(empty)'}",
        )
    base = entry.base_dir
    fallback = scan_fallback(base, max_depth=3)
    tree = build_folder_tree(base, max_depth=3)
    # Issue #313: scan_fallback が 1 件以上拾った場合は「候補なし」と矛盾するため、
    # suggest_patterns hit 経路と同じ「N 件候補あり」文言で統一する。
    # XlsxPickerDialog が候補一覧を出すのは fallback 結果も含めて全て同じため、
    # Treeview の詳細列と Dialog の表示を整合させる狙い。
    if fallback:
        message = f"{len(fallback)} 件候補あり、確認後に選択してください"
    else:
        message = "候補なし、フォルダから選択してください"
    return ResolveResult(
        status=CPlacementStatus.NEEDS_REVIEW,
        candidates=fallback,
        folder_tree=tree,
        message=message,
    )


def _normalize_name(name: str) -> str:
    """氏名比較用の正規化 (PR-γ v2: text_norm に統合)。

    PR-γ v2 まで: ``replace + strip`` のみで NFKC 欠落 (全角→半角効かず)。
    本関数は xlsx シート名 lookup に使われており、シート名は人手で作成されて
    全角/半角揺れが発生しやすい。``normalize_lookup_key`` への統合で NFKC を
    確実に通す。
    """
    return normalize_lookup_key(name)


def find_sheet_for_user(xlsx_path: Path, user_name: str) -> tuple[str | None, list[str]]:
    """xlsx 内のシート名から利用者氏名にマッチするものを探す。"""
    if not xlsx_path.exists():
        return None, []
    target = _normalize_name(user_name)
    with open(xlsx_path, "rb") as f:
        wb = load_workbook(io.BytesIO(f.read()), read_only=True)
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()
    matches = [n for n in names if _normalize_name(n) == target]
    if len(matches) == 1:
        return matches[0], names
    return None, names


def resolve_facility(facility_name: str, routing: Mapping[str, str]) -> str | None:
    """居宅名 → FAX フォルダ名 を解決する。

    PR-γ v1: 表記揺れ吸収のため lookup 時に ``normalize_lookup_key`` で正規化する。
    ``routing`` 側 (config.py の ``_load_checklist`` で読込時に正規化済み) と
    query 側の両方を同じ関数で正規化することで、全角/半角空白・全角/半角英数・
    括弧等の表記揺れを業務責任者の意識から外す。
    """
    key = normalize_lookup_key(facility_name)
    return routing.get(key)


def _resolve_chosen_staff(
    parsed_staffs: list[str],
    cfg: ChecklistConfig,
    year: int,
    month: int,
    result: CPlacementResult,
) -> str | None:
    """担当者複数 (Issue #314) の解決ロジック。chosen staff (元表記) を返す。

    解決順序 (Codex review High #3):
        1. 空 (row.staff="") → SKIPPED_NO_STAFF (result in-place 更新、None を返す)
        2. 単独 (len==1) → そのまま chosen として返す (既存単独経路)
        3. 複数 (len>=2):
            a. staff_choice_cache hit → cache value (normalize_lookup_key 形式、
               High #1) を parsed_staffs から元表記復元して返す
            b. cache miss + 全員未登録 → SKIPPED_NO_STAFF
            c. cache miss + 1 名以上登録済 → NEEDS_REVIEW_STAFF (登録済 1 名のみでも
               自動確定しない、High #4)。staff_candidates は元表記全員、message に
               部分 hit 時は "未登録あり: X, Y" を必ず含める (UI 表示契約)。

    Returns:
        chosen staff (元表記) — 後段の通常 xlsx 解決経路に進める単独担当者名
        None — result.status が SKIPPED_NO_STAFF / NEEDS_REVIEW_STAFF に設定済
    """
    if not parsed_staffs:
        result.status = CPlacementStatus.SKIPPED_NO_STAFF
        result.message = f"担当者マッピング未登録: {result.row.staff}"
        return None

    if len(parsed_staffs) == 1:
        return parsed_staffs[0]

    # 複数担当者: staff_choice_cache lookup
    key = staff_choice_cache_key(parsed_staffs, year, month)
    cached_normalized = cfg.staff_choice_cache.get(key)
    if cached_normalized is not None:
        # cache value は normalize_lookup_key 形式 (High #1)。parsed_staffs (元表記)
        # から normalize 一致する 1 件を復元して chosen として返す。
        for s in parsed_staffs:
            if normalize_lookup_key(s) == cached_normalized:
                return s
        # cache stale (parsed_staffs に含まれない正規化値が cache に残った場合)。
        # cache miss と同等の安全側に倒し、NEEDS_REVIEW_STAFF で人間判断を求める。
        logger.info(
            "staff_choice_cache stale for key=%s value=%s",
            key, cached_normalized,
        )

    # cache miss: 全員 mapping 登録済か判定
    registered: list[str] = []
    unregistered: list[str] = []
    for s in parsed_staffs:
        if normalize_lookup_key(s) in cfg.report_staff:
            registered.append(s)
        else:
            unregistered.append(s)

    if not registered:
        # 全員未登録 → 既存 SKIPPED_NO_STAFF 経路と整合
        result.status = CPlacementStatus.SKIPPED_NO_STAFF
        result.message = f"担当者マッピング未登録: {result.row.staff} (全員未登録)"
        return None

    # 1 名以上登録済 → NEEDS_REVIEW_STAFF (1 名のみでも自動確定しない、High #4)
    result.status = CPlacementStatus.NEEDS_REVIEW_STAFF
    result.staff_candidates = parsed_staffs  # 元表記全員 (UI で disable 制御)
    if unregistered:
        result.message = (
            f"{len(parsed_staffs)} 名中 {len(registered)} 名のみ登録済 "
            f"(未登録あり: {', '.join(unregistered)})、登録済から選択してください"
        )
    else:
        result.message = f"{len(parsed_staffs)} 名から担当者を選択してください"
    return None


def plan_c_placement(
    rows: list[ChecklistRow],
    cfg: ChecklistConfig,
    year: int,
    month: int,
) -> list[CPlacementResult]:
    """C 配置の計画を立てる（実 PDF 化はしない）。

    各行ごとに:
        1. 居宅 → FAX フォルダ resolve
        2. 担当者 → 単独/複数 を _resolve_chosen_staff で判定 (Issue #314)
        3. chosen staff → ReportStaffEntry resolve
        4. resolve_xlsx で cache hit / 候補抽出 / フォールバック
        5. PENDING のみシート検査して target_pdf 確定
        6. NEEDS_REVIEW / NEEDS_REVIEW_STAFF は UI に渡す
    """
    # Issue #27 続編 G Phase 3a: ChecklistConfig.fax_root は Path 型に移行済 (重複ラップ除去)。
    fax_root = cfg.fax_root
    results: list[CPlacementResult] = []
    for row in rows:
        result = CPlacementResult(row=row)
        fax_folder = resolve_facility(row.facility, cfg.facility_routing)
        if not fax_folder:
            result.status = CPlacementStatus.SKIPPED_NO_FACILITY
            result.message = f"居宅マッピング未登録: {row.facility}"
            results.append(result)
            continue

        # Issue #314: 担当者複数 (`/` `／` 区切り) の解決
        parsed_staffs = parse_multi_staff(row.staff)
        chosen_staff = _resolve_chosen_staff(
            parsed_staffs, cfg, year, month, result
        )
        if chosen_staff is None:
            # result.status は SKIPPED_NO_STAFF / NEEDS_REVIEW_STAFF に設定済
            results.append(result)
            continue

        # PR-γ v1: lookup 表記揺れ吸収（staff 側）。
        # message 表示は chosen_staff（業務責任者が分かる表記のまま）。
        staff_entry = cfg.report_staff.get(normalize_lookup_key(chosen_staff))
        if staff_entry is None:
            # 単独担当者で未登録、または cache stale 経由で復元失敗時の安全網
            result.status = CPlacementStatus.SKIPPED_NO_STAFF
            result.message = f"担当者マッピング未登録: {chosen_staff}"
            results.append(result)
            continue

        resolved = resolve_xlsx(
            staff=chosen_staff,
            entry=staff_entry,
            year=year,
            month=month,
            cache=cfg.xlsx_path_cache,
        )

        if resolved.status == CPlacementStatus.NEEDS_REVIEW:
            result.status = CPlacementStatus.NEEDS_REVIEW
            result.xlsx_candidates = resolved.candidates
            result.folder_tree = resolved.folder_tree
            result.message = resolved.message
            # NEEDS_REVIEW では target_pdf 確定不能、ユーザー選択後に再 plan
            results.append(result)
            continue

        if resolved.status == CPlacementStatus.SKIPPED_NO_XLSX:
            result.status = CPlacementStatus.SKIPPED_NO_XLSX
            result.message = resolved.message
            results.append(result)
            continue

        # PENDING（cache hit または legacy fallback）
        xlsx_path = resolved.xlsx_path
        assert xlsx_path is not None
        sheet_name, all_sheets = find_sheet_for_user(xlsx_path, row.name)
        if sheet_name is None:
            if all_sheets:
                result.sheet_candidates = all_sheets
            result.status = CPlacementStatus.SKIPPED_NO_SHEET
            result.message = f"利用者シート未発見: {row.name}"
            results.append(result)
            continue
        target = fax_root / fax_folder / cfg.c_output_subfolder / f"{row.name}.pdf"
        result.xlsx_path = xlsx_path
        result.sheet_name = sheet_name
        result.target_pdf = target
        result.status = CPlacementStatus.PENDING
        # PR (xlsx-visibility): cache hit / legacy 経由で確定した行は「自動:」prefix。
        # resolve_xlsx の legacy 経路は ``"自動: <name> (legacy)"`` を返すので尊重し、
        # cache hit 経路 (resolved.message 空) は basename を埋める。
        result.message = resolved.message or f"自動: {xlsx_path.name}"
        results.append(result)
    return results


def apply_xlsx_selection(
    result: CPlacementResult,
    xlsx_path: Path,
    cfg: ChecklistConfig,
) -> None:
    """ユーザーがレビュー UI で選択した xlsx_path を result に in-place 反映する。

    シート検査 → 利用者シート存在 + 単独 → PENDING + target_pdf 確定。
    シート未検出または曖昧 → SKIPPED_NO_SHEET / SKIPPED_AMBIGUOUS_SHEET。

    呼び出し側 (UI) は本関数の後に必要に応じて cfg.xlsx_path_cache を更新する。
    本関数は cache 操作を行わない（責務分離）。
    """
    # Issue #27 続編 G Phase 3a: ChecklistConfig.fax_root は Path 型に移行済 (重複ラップ除去)。
    fax_root = cfg.fax_root
    fax_folder = resolve_facility(result.row.facility, cfg.facility_routing)
    if not fax_folder:
        result.status = CPlacementStatus.SKIPPED_NO_FACILITY
        result.message = f"居宅マッピング未登録: {result.row.facility}"
        return
    sheet_name, all_sheets = find_sheet_for_user(xlsx_path, result.row.name)
    if sheet_name is None:
        if all_sheets:
            result.sheet_candidates = all_sheets
        result.status = CPlacementStatus.SKIPPED_NO_SHEET
        result.message = f"利用者シート未発見: {result.row.name}"
        return
    target = fax_root / fax_folder / cfg.c_output_subfolder / f"{result.row.name}.pdf"
    result.xlsx_path = xlsx_path
    result.sheet_name = sheet_name
    result.target_pdf = target
    result.status = CPlacementStatus.PENDING
    # NEEDS_REVIEW 時のフィールドはクリア（UI に古い候補を表示させない）
    result.xlsx_candidates = []
    result.folder_tree = None
    # PR (xlsx-visibility): 「選択: <basename>」prefix で「人間が選択した行」を可視化。
    # cache hit 経路の「自動: <basename>」と区別でき、業務責任者は配置確認モーダルを
    # 開かなくても Treeview 上で起源と対象ファイルを判別できる。
    result.message = f"選択: {xlsx_path.name}"


def execute_c_placement(
    results: list[CPlacementResult],
    exporter: ExcelExporter | None,
    log_dir: Path = Path(""),
    *,
    dry_run: bool = False,
) -> list[CPlacementResult]:
    """PENDING の plan を Excel COM で PDF 化して配置する。

    ``log_dir`` が指定されている場合、各行ごとに ``c_placement`` 監査ログに
    成功/失敗 record を JSON Lines で追記する（PR-α v3 の業務安全性層）。

    ``dry_run=True`` の場合は ExcelExporter を呼ばず PDF を実際に書き込まない
    （``exporter`` は ``None`` を渡してよい）。各行のパス解決 + シート検査が
    正しく完了したかだけを検証し、``status`` は PENDING のまま、``message`` に
    ``dry-run: 配置可能`` を記録する。監査ログには ``dry_run=true`` フラグを
    付与し、実配置とは区別できる。本番前の動作テスト用（PR-ζ v1）。
    """
    from wiseman_hub.audit import append_audit_record

    if not dry_run and exporter is None:
        raise ValueError("execute_c_placement: exporter must be provided when dry_run=False")
    try:
        for r in results:
            if r.status != CPlacementStatus.PENDING:
                continue
            if r.xlsx_path is None or r.sheet_name is None or r.target_pdf is None:
                r.status = CPlacementStatus.ERROR
                r.message = "internal: missing xlsx/sheet/target"
                continue
            if dry_run:
                # 実 PDF 書込なし。パス解決 + シート検査が plan_c_placement
                # で完了済みなので、ここでは「配置可能」と確認するだけ。
                # status は PENDING のまま、再実行（実配置）が可能な状態を保つ。
                # 「自動: <basename>」「選択: <basename>」起源 prefix が
                # plan_c_placement / apply_xlsx_selection で埋められているため、
                # dry-run 実行後も Treeview で起源を確認できるよう prev message を
                # 保持する。silent-failure H-4 / Codex Low 指摘対応: dry-run を
                # 複数回実行した時に "dry-run: 配置可能 (dry-run: 配置可能 (...))"
                # と入れ子化するのを防ぐため、既存 "dry-run:" prefix を検出して
                # idempotent に再ラップしない。
                prev = r.message
                if prev.startswith("dry-run:"):
                    pass  # 既に dry-run 表示 → 再ラップしない (idempotent)
                elif prev:
                    r.message = f"dry-run: 配置可能 ({prev})"
                else:
                    r.message = "dry-run: 配置可能"
            else:
                assert exporter is not None  # 上の guard で保証済（mypy 用）
                try:
                    exporter.export_first_page(
                        r.xlsx_path, r.sheet_name, r.target_pdf
                    )
                    # 二重ガード: exporter 側の存在確認に加えて、ここでも最終的な
                    # 物理ファイル存在を検証する（exporter 実装が黙って成功扱い
                    # する事案 / Excel COM サイレント失敗事案の再発防止）。
                    if not r.target_pdf.exists():
                        raise RuntimeError(
                            f"PDF was not written to: {r.target_pdf}"
                        )
                    if r.target_pdf.stat().st_size == 0:
                        raise RuntimeError(
                            f"PDF is empty (0 bytes): {r.target_pdf}"
                        )
                    r.status = CPlacementStatus.SUCCESS
                except Exception as exc:  # MVP: 詳細分類は後段
                    r.status = CPlacementStatus.ERROR
                    r.message = (
                        f"export failed: {exc.__class__.__name__}: {exc}"
                    )
                    logger.exception("Excel export failed for %s", r.row.name)
            append_audit_record(
                log_dir,
                kind="c_placement",
                record={
                    "user": r.row.name,
                    "facility": r.row.facility,
                    "staff": r.row.staff,
                    "xlsx_path": str(r.xlsx_path) if r.xlsx_path else None,
                    "sheet_name": r.sheet_name,
                    "target_pdf": (
                        str(r.target_pdf) if r.target_pdf else None
                    ),
                    "status": r.status.value,
                    "message": r.message,
                    "dry_run": dry_run,
                },
            )
    finally:
        if not dry_run and exporter is not None:
            exporter.close()
    return results
