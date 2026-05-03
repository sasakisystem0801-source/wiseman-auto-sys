"""居宅名 → FAX 事業所フォルダ名 の対照表ドラフトを生成する一回限りスクリプト。

入力:
    - スプレッドシート（居宅名 60 件）: /tmp/wiseman-sheet/sheet.xlsx
    - FAX フォルダ名 40 件: 本スクリプト内ハードコード（Session 39 取得値）

出力:
    docs/handoff/facility-mapping-draft.md（confidence 別表）

実行:
    uv run python scripts/draft_facility_mapping.py
"""

from __future__ import annotations

import re
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

SHEET_PATH = Path("/tmp/wiseman-sheet/sheet.xlsx")
OUT_PATH = Path("docs/handoff/facility-mapping-draft.md")

FAX_FOLDERS: list[str] = [
    "LEBEN(メール)",
    "RIN(メール)",
    "あおぞら(FAX)",
    "あゆみ愛(FAX)",
    "あん(メール)",
    "きなり(メール)※持参",
    "なごみの里(メール)",
    "のもと本店（メール）",
    "ほおずき（FAX）",
    "まほろば(メール)",
    "むれさき(FAX)",
    "やまさん家（メール）",
    "アミー(FAX)",
    "オレンジ（メール）",
    "ケアプラン太子（メール）※持参",
    "ケアプラン正條（FAX）",
    "ケアプラン笑楽西明石（メール）",
    "シスナブ御津(メール)",
    "シルバーケア(メール)",
    "スマイル(メール)※持参",
    "ツカザキ あぼし(メール)",
    "ツカザキ 広畑(メール)",
    "フラワー居宅（メール）",
    "メディカルプラン結(FAX)",
    "リリーたつの(メール)坂川CM",
    "リリーライフ(メール)西久保CM",
    "大津みやびの(メール)",
    "大津地域包括支援センター（メール）",
    "太子の郷（FAX）※持参",
    "太子町地域包括（メール）※持参",
    "太子病院(メール)",
    "姫路・勝原ホーム(メール)",
    "姫路医療生活協同組合 あぼし(メール)",
    "広畑地域包括（メール）",
    "朝日地域包括(メール)",
    "清住園(FAX)",
    "緑ヶ丘(FAX)※17時まで",
    "花の里(メール)",
    "西はりま(FAX)",
    "銀の櫂（メール）",
]


def normalize_core(name: str) -> str:
    """送付注記・組織形態 suffix を除去して比較用 core を作る。"""
    s = unicodedata.normalize("NFKC", name)
    # 送付方法注記
    s = re.sub(r"[(](メール|FAX|fax)[)]", "", s)
    s = re.sub(r"※[^\s　]*", "", s)
    # 担当 CM 注記
    s = re.sub(r"\s*(坂川CM|西久保CM)\s*", "", s)
    # 組織形態
    s = s.replace("居宅介護支援事業所", "")
    s = s.replace("地域包括支援センター", "地域包括")
    s = s.replace("サポートセンター", "")
    s = s.replace("ケアプランセンター", "")
    s = s.replace("有限会社", "")
    s = s.replace("姫路市", "")
    # 中黒・空白
    s = s.replace("　", "").replace(" ", "")
    s = s.replace("・", "")
    return s.strip()


def score(home_core: str, fax_core: str) -> tuple[float, str]:
    """home と fax の core 同士の類似度スコア + マッチタイプ。"""
    if not home_core or not fax_core:
        return 0.0, "empty"
    if home_core == fax_core:
        return 1.0, "exact"
    if home_core in fax_core:
        return len(home_core) / len(fax_core), "home_in_fax"
    if fax_core in home_core:
        return len(fax_core) / len(home_core), "fax_in_home"
    # 共通文字数 / max
    common = sum((Counter(home_core) & Counter(fax_core)).values())
    if common >= 2:
        return (common / max(len(home_core), len(fax_core))) * 0.6, "char_overlap"
    return 0.0, "none"


def confidence_of(top1: float, top2: float) -> str:
    if top1 >= 0.8 and (top1 - top2) >= 0.2:
        return "high"
    if top1 >= 0.6:
        return "medium"
    if top1 >= 0.4:
        return "low"
    return "none"


def extract_homes() -> list[tuple[str, int]]:
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    counter: Counter[str] = Counter()
    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = [c.value for c in ws[1]]
        target_col = 14
        for i, h in enumerate(headers):
            if h and ("居宅" in str(h) or "事業所" in str(h)):
                target_col = i
                break
        for row in ws.iter_rows(min_row=2, values_only=True):
            if target_col < len(row) and row[target_col]:
                v = str(row[target_col]).strip()
                if v:
                    counter[v] += 1
    return counter.most_common()


def main() -> None:
    homes = extract_homes()
    fax_pairs = [(f, normalize_core(f)) for f in FAX_FOLDERS]

    high: list[tuple[str, int, str, float]] = []
    medium: list[tuple[str, int, list[tuple[str, float]]]] = []
    low: list[tuple[str, int, list[tuple[str, float]]]] = []
    unmatched: list[tuple[str, int]] = []

    for home_name, occ in homes:
        home_core = normalize_core(home_name)
        scored: list[tuple[str, float, str]] = []
        for fax_name, fax_core in fax_pairs:
            sc, kind = score(home_core, fax_core)
            if sc > 0:
                scored.append((fax_name, sc, kind))
        scored.sort(key=lambda x: -x[1])
        if not scored:
            unmatched.append((home_name, occ))
            continue
        top1 = scored[0][1]
        top2 = scored[1][1] if len(scored) > 1 else 0.0
        conf = confidence_of(top1, top2)
        if conf == "high":
            high.append((home_name, occ, scored[0][0], top1))
        elif conf == "medium":
            medium.append((home_name, occ, [(s[0], s[1]) for s in scored[:3]]))
        elif conf == "low":
            low.append((home_name, occ, [(s[0], s[1]) for s in scored[:3]]))
        else:
            unmatched.append((home_name, occ))

    lines: list[str] = []
    lines.append("# 居宅 → FAX 事業所フォルダ 対照表ドラフト")
    lines.append("")
    lines.append(
        f"**生成元**: 60 居宅 × 40 FAX フォルダの自動マッチング（rule-based、AI 不使用）"
    )
    lines.append(
        "**目的**: ユーザーが目視で精査・確定するためのたたき台。修正後 JSON/TOML 化してアプリへ反映。"
    )
    lines.append("")
    lines.append(
        f"集計: HIGH {len(high)} / MEDIUM {len(medium)} / LOW {len(low)} / UNMATCHED {len(unmatched)}"
    )
    lines.append("")
    lines.append(
        "判定軸: `core` 同士（送付注記・組織形態 suffix を除去した本体）の包含・共通文字率。"
        "HIGH = top1 ≥ 0.8 かつ top2 と 0.2+ 差。MEDIUM = top1 ≥ 0.6。LOW = top1 ≥ 0.4。"
    )
    lines.append("")

    lines.append("## ✅ HIGH（自動採用候補、確認のみ）")
    lines.append("")
    lines.append("| 居宅名 | 出現 | → FAX フォルダ | score |")
    lines.append("|--------|------|----------------|-------|")
    for home, occ, fax, sc in high:
        lines.append(f"| {home} | {occ} | {fax} | {sc:.2f} |")
    lines.append("")

    lines.append("## ⚠ MEDIUM（要選択：top3 候補を提示）")
    lines.append("")
    lines.append("| 居宅名 | 出現 | 候補1 (score) | 候補2 | 候補3 |")
    lines.append("|--------|------|---------------|-------|-------|")
    for home, occ, cands in medium:
        c1 = f"{cands[0][0]} ({cands[0][1]:.2f})" if len(cands) > 0 else ""
        c2 = f"{cands[1][0]} ({cands[1][1]:.2f})" if len(cands) > 1 else ""
        c3 = f"{cands[2][0]} ({cands[2][1]:.2f})" if len(cands) > 2 else ""
        lines.append(f"| {home} | {occ} | {c1} | {c2} | {c3} |")
    lines.append("")

    lines.append("## ⚠ LOW（要手動：候補弱い）")
    lines.append("")
    lines.append("| 居宅名 | 出現 | 候補1 (score) | 候補2 | 候補3 |")
    lines.append("|--------|------|---------------|-------|-------|")
    for home, occ, cands in low:
        c1 = f"{cands[0][0]} ({cands[0][1]:.2f})" if len(cands) > 0 else ""
        c2 = f"{cands[1][0]} ({cands[1][1]:.2f})" if len(cands) > 1 else ""
        c3 = f"{cands[2][0]} ({cands[2][1]:.2f})" if len(cands) > 2 else ""
        lines.append(f"| {home} | {occ} | {c1} | {c2} | {c3} |")
    lines.append("")

    lines.append("## ❌ UNMATCHED（候補なし、手動入力必要）")
    lines.append("")
    lines.append("| 居宅名 | 出現 |")
    lines.append("|--------|------|")
    for home, occ in unmatched:
        lines.append(f"| {home} | {occ} |")
    lines.append("")

    lines.append("## 参考: 全 FAX フォルダ 40 件")
    lines.append("")
    for f in FAX_FOLDERS:
        lines.append(f"- {f}")
    lines.append("")

    OUT_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT_PATH} ({len(lines)} lines)")
    print(f"HIGH={len(high)}, MEDIUM={len(medium)}, LOW={len(low)}, UNMATCHED={len(unmatched)}")


if __name__ == "__main__":
    main()
